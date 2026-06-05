import pandas as pd
from pathlib import Path

# Rutas
ruta_caso1 = Path(r"C:\Users\Daniel_Dedalus\Desktop\DASP\DASP\patrones de concurrencia\1. Sección Crítica - Fiabilidad de Hilos\CriticalSectionSTExercise\logAnalisisSafeThread.csv")
ruta_caso2 = Path(r"C:\Users\Daniel_Dedalus\Desktop\DASP\DASP\patrones de concurrencia\1. Sección Crítica - Inicialización Temprana\CriticalSectionEarlyInitExercise\logAnalisisSafeThread.csv")

# Leer
df1 = pd.read_csv(ruta_caso1) if ruta_caso1.exists() else pd.DataFrame()
df2 = pd.read_csv(ruta_caso2) if ruta_caso2.exists() else pd.DataFrame()

# Crear Excel
with pd.ExcelWriter("comparativa_casos.xlsx", engine='openpyxl') as writer:
    # Primera hoja: datos uno al lado del otro
    hoja = writer.book.create_sheet("Comparativa")
    
    # Títulos de cada caso
    hoja['A1'] = "CASO 1: Sección Crítica - Fiabilidad de Hilos"
    hoja['G1'] = "CASO 2: Inicialización Temprana"
    
    # Escribir datos caso 1 a partir de fila 3, columna A
    if not df1.empty:
        for c, col in enumerate(df1.columns):
            hoja.cell(row=3, column=c+1, value=col)
        for i, row in df1.iterrows():
            for c, val in enumerate(row):
                hoja.cell(row=i+4, column=c+1, value=val)
    
    # Escribir datos caso 2 a partir de fila 3, columna G (columna 7)
    if not df2.empty:
        for c, col in enumerate(df2.columns):
            hoja.cell(row=3, column=c+7, value=col)
        for i, row in df2.iterrows():
            for c, val in enumerate(row):
                hoja.cell(row=i+4, column=c+7, value=val)
    
    # Ajustar ancho de columnas
    from openpyxl.utils import get_column_letter
    for col in range(1, 20):  # hasta columna T
        letra = get_column_letter(col)
        hoja.column_dimensions[letra].width = 15

print(f"✓ Archivo Excel generado: comparativa_casos.xlsx")